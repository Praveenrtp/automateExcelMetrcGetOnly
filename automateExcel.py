import requests
import pandas as pd
import numpy as np 
import openpyxl
import json



try:
    vendor_key = 'KXvRFsRK2OnTg6M63eXMFvmgCD235ultHrYfaeySDfOtAts5'
    user_api_key = 'FusVbe4Yv6W1DGNuxKNhByXU6RO6jSUPcbRCoRDD98VNXc4D'
    sandbox_url= 'https://sandbox-api-ca.metrc.com/'
    params = {'licenseNumber': 'C12-1000003-LIC', 'lastModifiedStart': '2021-08-04'}
    licenseNumber = 'C12-1000003-LIC'
    name = 'Cannabis - Microbusiness License'

    ignore = 'The follwing may Not be supported in the state you are testing for. Pleae refer to the The Metrc Web API Document page that can be found: https://api-XX.metrc.com/Documentation#getting-started <Change the XX for the abbreviated state for the corresponding page>'
    ignore_1 = 'GET /transfers/v1/{id}/deliveries                                                    This ID number is the same as the ID number at the top of a GET /transfers/v1/incoming, outgoing or rejected'
    ignore_2 = 'GET /transfers/v1/delivery/{id}/packages                                       This ID number is the Delivery ID number from a                          GET /transfers/v1/incoming, outgoing or rejected OR the first number returned by the                                                                  GET /transfers/v1/{id}/deliveries   '
    ignore_3 = 'GET /transfers/v1/delivery/{id}/packages/wholesale    This ID number is the Delivery ID number from a                          GET /transfers/v1/incoming, outgoing or rejected OR the first number returned by the GET /transfers/v1/{id}/deliveries '
    
    def multipleReplace(text, word):
        if 'id' in text:
            text = text.replace('{id}', word)
        if 'label' in text:
            text = text.replace('{label}', word)
        return text

    def get_response(req_url, original_string, req_params):
        if 'id' in original_string and 'plants' not in original_string and 'sales' not in original_string and 'transfers' not in original_string :
            g = multipleReplace(original_string, 'active')
            req_url = req_url+g
            r = requests.get(req_url, auth=(vendor_key, user_api_key), params = req_params)
            s = json.loads(r.text)
            val = s[0]['Id']
            return val

        if 'id' in original_string and 'transfers' in original_string:
            url = sandbox_url + 'transfers/v1/rejected'
            r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
            s = json.loads(r.text)
            val = s[0]['Id']
            return val

        if 'plants' in original_string:
            g = multipleReplace(original_string, 'inactive')
            req_url = req_url+g
            r = requests.get(req_url, auth=(vendor_key, user_api_key), params = req_params)
            s = json.loads(r.text)
            val = s[0]['Label']
            return val

        if 'packages' in original_string :
            g = multipleReplace(original_string, 'active')
            req_url = req_url+g
            r = requests.get(req_url, auth=(vendor_key, user_api_key), params = req_params)
            s = json.loads(r.text)
            val = s[0]['Label']
            return val


    source = openpyxl.load_workbook('GET_Only_Evaluation_2020_V1001.xlsx', read_only=False, keep_vba=True)
    sheetnames=['Admin', 'Cultivation', 'Packages', 'Lab', 'Sales', 'Transfers']
    parameter_holders=['id', 'plantbatches', 'plants', 'harvests']

    replace_format = {
        'id' : 'active',
        'label': 'active',
        'plants' : 'inactive',
    }

    xls = pd.ExcelFile('GET_Only_Evaluation_2020_V1001.xlsx')
    df1 = pd.read_excel(xls, 'Company Information')
    df1 = df1.replace(np.nan, '', regex=True)
    sheetname = source.get_sheet_by_name('Company Information')
    sheetname.cell(row=7,column=3).value = 'Lendica Corp'
    sheetname.cell(row=8, column=3).value = 'golendica.com'
    sheetname.cell(row=9,column=3).value = '6172862390'
    sheetname.cell(row=10, column=3).value = 'jared@golendica.com'
    sheetname.cell(row=11,column=3).value = '201 Washington Street Suite 2600'
    sheetname.cell(row=12, column=3).value = 'Boston, MA'
    sheetname.cell(row=13,column=3).value = '2108'
    sheetname.cell(row=14, column=3).value = 'Jared Shulman'
    sheetname.cell(row=15,column=3).value = 'jared@golendica.com'
    sheetname.cell(row=16, column=3).value = '6172862390'
    sheetname.cell(row=17,column=3).value = 'Jerry Shu'
    sheetname.cell(row=18, column=3).value = 'jerry@golendica.com'
    sheetname.cell(row=19,column=3).value = '6174595751'
    sheetname.cell(row=20, column=3).value = 'Lendica Access'
    sheetname.cell(row=21,column=3).value = vendor_key
    sheetname.cell(row=22, column=3).value = user_api_key
    print('went in')

    for i in sheetnames:
        df1 = pd.read_excel(xls, i)
        df1 = df1.replace(np.nan, '', regex=True)
        sheetname = source.get_sheet_by_name(i)
        if i == 'Lab':
            i = 'Lab Results'
        if i == 'Transfers':
            i = 'GET Transfers and Wholesale'
        df1.columns = df1.columns.str.lstrip()
        df1.columns = df1.columns.str.rstrip()
        for index,value in df1.iterrows():
            if 'Metrc Use Only' in value[i]:
                break

            if ignore in value[i]:
                break

            if 'GET' in value[i]:
                d = '/'.join(value[i].split('/')[1:])
                if any(check in d for check in replace_format):
                    url=sandbox_url
                    response_val = get_response(url, d, params)
                    dd = multipleReplace(d, str(response_val))
                    url=sandbox_url+'{}'.format(dd)
                    r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
                else:
                    url=sandbox_url+'{}'.format(d)
                    r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
                    

                if 'labtests' in value[i] :
                    params = {'licenseNumber': 'C12-1000003-LIC', 'lastModifiedStart': '2021-08-04'}
                    url = sandbox_url + 'packages/v1/active'
                    r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
                    s = json.loads(r.text)
                    val = s[0]['Id']
                    req_params = {'licenseNumber': 'C12-1000003-LIC','packageId': val}
                    new_url = sandbox_url + 'labtests/v1/results'
                    r = requests.get(new_url, auth=(vendor_key, user_api_key), params = req_params)
                    

                if 'sales' in value[i] :
                    new_url = sandbox_url + 'sales/v1/receipts/active'
                    r = requests.get(new_url, auth=(vendor_key, user_api_key), params = params)
                    

                if 'GET /sales/v1/receipts/{id}' in value[i]:
                    sheetname.cell(row=index+2,column=2).value = r.status_code
                    sheetname.cell(row=index+2, column=3).value = licenseNumber
                    sheetname.cell(row=index+2, column=4).value = 'N/A'
                    sheetname.cell(row=index+2,column=5).value = params['lastModifiedStart']
                    sheetname.cell(row=index+2, column=6).value = 'N/A'
                    sheetname.cell(row=index+2,column=7).value = sandbox_url + 'sales/v1/receipts/{id}'
                    sheetname.cell(row=index+2,column=8).value = '[]'

                if 'rejected' in value[i]:
                    new_url = sandbox_url + 'transfers/v1/rejected'
                    r = requests.get(new_url, auth=(vendor_key, user_api_key), params = params)
                    s = json.loads(r.text)
                    val = s[0]['Id']

                if ignore_1 in value[i]:
                    url = sandbox_url + 'transfers/v1/rejected'
                    r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
                    s = json.loads(r.text)
                    val = s[0]['Id']
                    d =  ignore_1.split(' ')
                    dd = d[1]
                    replace = multipleReplace(dd, str(val))
                    new_url = sandbox_url+'{}'.format(replace)
                    r = requests.get(new_url, auth=(vendor_key, user_api_key))

                if ignore_2 in value[i]:
                    url = sandbox_url + 'transfers/v1/rejected'
                    r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
                    s = json.loads(r.text)
                    val = s[0]['Id']
                    d =  ignore_1.split(' ')
                    dd = d[1]
                    replace = multipleReplace(dd, str(val))
                    new_url=sandbox_url+'{}'.format(replace)
                    r = requests.get(new_url, auth=(vendor_key, user_api_key))

                if ignore_3 in value[i]:
                    url = sandbox_url + 'transfers/v1/rejected'
                    r = requests.get(url, auth=(vendor_key, user_api_key), params = params)
                    s = json.loads(r.text)
                    val = s[0]['Id']
                    d =  ignore_1.split(' ')
                    dd = d[1]
                    replace = multipleReplace(dd, str(val))
                    new_url=sandbox_url+'{}'.format(replace)
                    r = requests.get(new_url, auth=(vendor_key, user_api_key))

                if i == 'Admin':
                    sheetname.cell(row=index+2,column=5).value = name
                    sheetname.cell(row=index+2,column=6).value = url
                    sheetname.cell(row=index+2, column=7).value = r.text
                else:              
                    sheetname.cell(row=index+2,column=5).value = params['lastModifiedStart']
                    sheetname.cell(row=index+2, column=6).value = 'N/A'
                    sheetname.cell(row=index+2,column=7).value = url

                if 'id' in value[i]:
                    sheetname.cell(row=index+2, column=4).value = response_val
                else :  
                    sheetname.cell(row=index+2, column=4).value = 'N/A'
                    
                sheetname.cell(row=index+2, column=8).value = r.text
                sheetname.cell(row=index+2,column=2).value = r.status_code
                sheetname.cell(row=index+2, column=3).value = licenseNumber
                source.save('GET_Only_Evaluation_2020_V1001.xlsx')
                print('success')
            else:
                continue
except (Exception):
    pass
    



